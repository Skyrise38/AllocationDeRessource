import openpyxl
from openpyxl import load_workbook
import random
import munkres
from munkres import Munkres, print_matrix

nb_choix_projet=3

def verifier_tableau(fichier_excel,nb_choix_projet):
    wb = load_workbook(fichier_excel)
    feuille = wb.active
    nb_lignes = feuille.max_row
    nb_colonnes = feuille.max_column
    for i in range(2,nb_lignes):
        somme=0
        tab=[]
        for j in range(2,nb_colonnes):
            if feuille.cell(column=j, row=i).value!=None:
                for caractere in str(feuille.cell(column=j, row=i).value):
                    if caractere.isdigit()==False:
                        print("Erreur, tableau invalide")
                        return False
                if int(feuille.cell(column=j, row=i).value)>nb_choix_projet:
                    print("Erreur, tableau invalide")
                    return False
                somme+=int(feuille.cell(column=j, row=i).value)
                tab.append(int(feuille.cell(column=j, row=i).value))
        if somme!=nb_choix_projet*(nb_choix_projet+1)/2:
            print("Erreur, tableau invalide")
            return False
        if len(tab)!=nb_choix_projet:
            print("Erreur, tableau invalide")
            return False
        for k in range(1,nb_choix_projet+1):
            if k not in tab:
                print("Erreur, tableau invalide")
                return False
    print("Tableau valide")
    return True
    
verifier_tableau("sujets choisis test.xlsx",nb_choix_projet)

