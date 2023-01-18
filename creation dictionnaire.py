import openpyxl
from openpyxl import load_workbook

def creation_dictionnaire(fichier_excel):
    wb = load_workbook(fichier_excel)
    feuille = wb.active
    nb_lignes = feuille.max_row
    nb_colonnes = feuille.max_column
    eleves={}
    tab_valeurs = ["1","2","3"]
    for i in range(2,nb_lignes):
        tab_choix = []
        for j in range(2,nb_colonnes):
            if (str(feuille.cell(column=j, row=i).value)) not in tab_valeurs:
                tab_choix.append("10")
            else :
                tab_choix.append(feuille.cell(column=j, row=i).value)
        eleves[i]={}
        eleves[i]["Nom"]=feuille.cell(column=1, row=i).value
        eleves[i]["tab_choix"]=tab_choix

    print(eleves)
    print(2)

creation_dictionnaire("sujets choisis.xlsx")